import io

import openpyxl

from app.models import Game, Season, Team
from tests.conftest import auth_headers, make_league, make_user

_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_workbook(rows: list[tuple]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "schedule"
    ws.append(["match_no", "date", "time", "venue", "away_team", "home_team"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _setup(db_session, suffix: str):
    league = make_league(db_session, name=f"Sched {suffix}", slug=f"sched-{suffix.lower()}")
    admin = make_user(
        db_session, email=f"admin@sched-{suffix.lower()}.test", role="admin", league_id=league.id,
    )
    season = Season(league_id=league.id, year=2026, name="2026")
    away = Team(league_id=league.id, name=f"{suffix} Away")
    home = Team(league_id=league.id, name=f"{suffix} Home")
    db_session.add_all([season, away, home])
    db_session.flush()
    return league, admin, season, away, home


def _upload(client, season_id, content, headers, **params):
    return client.post(
        f"/api/v1/seasons/{season_id}/games/import",
        files={"file": ("schedule.xlsx", content, _XLSX_CONTENT_TYPE)},
        headers=headers,
        params=params,
    )


def test_template_download(client, db_session) -> None:
    _, admin, season, _away, _home = _setup(db_session, "A")

    resp = client.get(
        f"/api/v1/seasons/{season.id}/games/template.xlsx", headers=auth_headers(admin),
    )
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb["schedule"].max_row == 1


def test_import_requires_confirm_then_commits(client, db_session) -> None:
    _, admin, season, away, home = _setup(db_session, "B")
    content = _build_workbook([(1, "2026-08-01", "18:00", "示範球場", away.name, home.name)])

    preview = _upload(client, season.id, content, auth_headers(admin))
    assert preview.status_code == 200
    body = preview.json()
    assert body["committed"] is False
    assert body["valid_rows"] == 1
    assert db_session.query(Game).filter(Game.season_id == season.id).count() == 0

    committed = _upload(client, season.id, content, auth_headers(admin), confirm="true")
    assert committed.status_code == 200
    assert committed.json()["committed"] is True

    games = db_session.query(Game).filter(Game.season_id == season.id).all()
    assert len(games) == 1
    assert games[0].away_team_id == away.id
    assert games[0].home_team_id == home.id
    assert games[0].code == "1"
    assert games[0].venue == "示範球場"


def test_import_unknown_team_name_is_row_error(client, db_session) -> None:
    _, admin, season, away, _home = _setup(db_session, "C")
    content = _build_workbook([(1, "2026-08-01", "", "", away.name, "不存在的隊伍")])

    resp = _upload(client, season.id, content, auth_headers(admin), confirm="true")
    assert resp.status_code == 200
    body = resp.json()
    assert body["committed"] is False
    assert any(e["field"] == "home_team" for e in body["errors"])
    assert db_session.query(Game).filter(Game.season_id == season.id).count() == 0


def test_import_same_team_twice_is_row_error(client, db_session) -> None:
    _, admin, season, away, _home = _setup(db_session, "D")
    content = _build_workbook([(1, "2026-08-01", "", "", away.name, away.name)])

    resp = _upload(client, season.id, content, auth_headers(admin))
    assert resp.status_code == 200
    assert resp.json()["committed"] is False


def test_power_cannot_import_schedule(client, db_session) -> None:
    league, _admin, season, away, home = _setup(db_session, "E")
    power = make_user(db_session, email="power@sched-e.test", role="power", league_id=league.id)
    content = _build_workbook([(1, "2026-08-01", "", "", away.name, home.name)])

    resp = _upload(client, season.id, content, auth_headers(power))
    assert resp.status_code == 403
