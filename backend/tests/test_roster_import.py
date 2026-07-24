import io

import openpyxl

from app.models import Player, Team
from tests.conftest import auth_headers, make_league, make_user

_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_workbook(rows: list[tuple]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "roster"
    ws.append(["number", "name", "positions", "bats_throws", "email", "phone", "birthdate"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _setup_admin_team(db_session, suffix: str):
    league = make_league(db_session, name=f"Import {suffix}", slug=f"import-{suffix.lower()}")
    admin = make_user(
        db_session, email=f"admin@import-{suffix.lower()}.test", role="admin",
        league_id=league.id,
    )
    team = Team(league_id=league.id, name=f"{suffix} Team")
    db_session.add(team)
    db_session.flush()
    return league, admin, team


def _upload(client, team_id, content, headers, **params):
    return client.post(
        f"/api/v1/teams/{team_id}/roster/import",
        files={"file": ("roster.xlsx", content, _XLSX_CONTENT_TYPE)},
        headers=headers,
        params=params,
    )


def test_import_valid_9_row_xlsx_creates_9_players(client, db_session) -> None:
    _, admin, team = _setup_admin_team(db_session, "A")
    rows = [(i, f"Player {i}", "OF", "R/R", "", "", "") for i in range(1, 10)]

    resp = _upload(client, team.id, _build_workbook(rows), auth_headers(admin))
    assert resp.status_code == 200
    body = resp.json()
    assert body["valid_rows"] == 9
    assert body["errors"] == []
    assert body["committed"] is True

    count = db_session.query(Player).filter(Player.team_id == team.id).count()
    assert count == 9


def test_import_with_duplicate_number_rejects_nothing_committed(client, db_session) -> None:
    _, admin, team = _setup_admin_team(db_session, "B")
    rows = [
        (1, "Player One", "OF", "R/R", "", "", ""),
        (1, "Player Two", "1B", "L/L", "", "", ""),
    ]

    resp = _upload(client, team.id, _build_workbook(rows), auth_headers(admin))
    assert resp.status_code == 200
    body = resp.json()
    assert body["committed"] is False
    assert any(e["field"] == "number" for e in body["errors"])

    count = db_session.query(Player).filter(Player.team_id == team.id).count()
    assert count == 0


def test_template_download_round_trips_with_import(client, db_session) -> None:
    _, admin, team = _setup_admin_team(db_session, "C")

    tmpl_resp = client.get(
        f"/api/v1/teams/{team.id}/roster/template.xlsx", headers=auth_headers(admin),
    )
    assert tmpl_resp.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(tmpl_resp.content))
    ws = wb["roster"]
    ws.append([1, "Round Trip Player", "SS", "R/R", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)

    resp = _upload(client, team.id, buf.getvalue(), auth_headers(admin))
    assert resp.status_code == 200
    assert resp.json()["committed"] is True
    assert resp.json()["valid_rows"] == 1


def test_replace_mode_requires_confirm(client, db_session) -> None:
    league, admin, team = _setup_admin_team(db_session, "D")
    existing = Player(league_id=league.id, team_id=team.id, name="Existing", number=5)
    db_session.add(existing)
    db_session.flush()

    content = _build_workbook([(1, "New Player", "OF", "R/R", "", "", "")])

    preview = _upload(client, team.id, content, auth_headers(admin), mode="replace")
    assert preview.status_code == 200
    assert preview.json()["committed"] is False
    active = db_session.query(Player).filter(Player.team_id == team.id, Player.status == "active")
    assert active.count() == 1

    committed = _upload(
        client, team.id, content, auth_headers(admin), mode="replace", confirm="true",
    )
    assert committed.status_code == 200
    assert committed.json()["committed"] is True

    db_session.refresh(existing)
    assert existing.status == "left"
    active_players = active.all()
    assert len(active_players) == 1
    assert active_players[0].name == "New Player"


def test_power_cannot_import_roster(client, db_session) -> None:
    league, admin, team = _setup_admin_team(db_session, "E")
    power = make_user(db_session, email="power@import-e.test", role="power", league_id=league.id)
    content = _build_workbook([(1, "X", None, None, "", "", "")])

    resp = _upload(client, team.id, content, auth_headers(power))
    assert resp.status_code == 403


def _build_full_workbook(rows: list[tuple]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "roster"
    ws.append(
        ["number", "name", "positions", "bats_throws", "title", "birthdate", "national_id", "email", "phone"],
    )
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_import_full_fields_round_trip_and_national_id_hidden_from_list(client, db_session) -> None:
    _, admin, team = _setup_admin_team(db_session, "F")
    rows = [
        (1, "Full Fields", "OF", "R/R", "領隊", "2000-01-15", "A123456789", "full@example.com", "0912345678"),
        (2, "Default Title", "1B", None, "", "", "", "", ""),
    ]

    resp = _upload(client, team.id, _build_full_workbook(rows), auth_headers(admin))
    assert resp.status_code == 200
    assert resp.json()["committed"] is True

    players = {p.number: p for p in db_session.query(Player).filter(Player.team_id == team.id)}
    assert players[1].title == "manager"
    assert str(players[1].birthdate) == "2000-01-15"
    assert players[1].national_id == "A123456789"
    assert players[1].email == "full@example.com"
    assert players[1].phone == "0912345678"
    assert players[2].title == "member"  # blank title defaults to member

    list_resp = client.get(f"/api/v1/teams/{team.id}/players", headers=auth_headers(admin))
    assert list_resp.status_code == 200
    for row in list_resp.json():
        assert "national_id" not in row

    detail_resp = client.get(f"/api/v1/players/{players[1].id}", headers=auth_headers(admin))
    assert detail_resp.status_code == 200
    assert detail_resp.json()["national_id"] == "A123456789"


def test_import_invalid_title_and_national_id_rejected(client, db_session) -> None:
    _, admin, team = _setup_admin_team(db_session, "G")
    rows = [(1, "Bad Row", None, None, "隊醫", "", "12345", "", "")]

    resp = _upload(client, team.id, _build_full_workbook(rows), auth_headers(admin))
    assert resp.status_code == 200
    body = resp.json()
    assert body["committed"] is False
    fields = {e["field"] for e in body["errors"]}
    assert "title" in fields
    assert "national_id" in fields
