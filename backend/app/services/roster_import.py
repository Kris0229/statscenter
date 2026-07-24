"""Excel roster import: parsing + row validation only (BUILD_SPEC §5).

Committing parsed rows to the database is the API layer's job
(app/api/roster_import.py), so this stays a pure, easily-testable function.
"""
import io
import re
from dataclasses import dataclass
from datetime import date, datetime

import openpyxl

TEMPLATE_COLUMNS = [
    "number", "name", "positions", "bats_throws",
    "title", "birthdate", "national_id", "email", "phone",
]

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_BATS_VALUES = {"R", "L", "S"}
_THROWS_VALUES = {"R", "L"}
# Taiwan national ID: 1 letter + 9 digits.
_NATIONAL_ID_RE = re.compile(r"^[A-Za-z]\d{9}$")
_TITLE_LABELS = {"領隊": "manager", "教練": "coach", "隊長": "captain", "隊員": "member"}
_TITLE_VALUES = set(_TITLE_LABELS.values())


@dataclass
class RosterRow:
    row: int
    number: int
    name: str
    positions: str | None
    bats: str | None
    throws: str | None
    title: str
    birthdate: date | None
    national_id: str | None
    email: str | None
    phone: str | None


@dataclass
class RowError:
    row: int
    field: str
    msg: str


@dataclass
class ParseResult:
    valid_rows: list[RosterRow]
    errors: list[RowError]


def build_template_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "roster"
    ws.append(TEMPLATE_COLUMNS)
    return wb


def _blank(value: object) -> bool:
    return value is None or str(value).strip() == ""


def _parse_birthdate(raw: object) -> tuple[date | None, str | None]:
    """Returns (value, error_msg). Accepts Excel date cells (already a date/
    datetime object) or an ISO 'YYYY-MM-DD' string."""
    if isinstance(raw, datetime):
        return raw.date(), None
    if isinstance(raw, date):
        return raw, None
    try:
        return date.fromisoformat(str(raw).strip()), None
    except ValueError:
        return None, "expected a date cell or 'YYYY-MM-DD'"


def parse_roster_workbook(content: bytes) -> ParseResult:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb["roster"] if "roster" in wb.sheetnames else wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1), ())
    header = [str(c.value).strip().lower() if c.value is not None else "" for c in header_row]
    col_index = {name: i for i, name in enumerate(header)}

    missing = [c for c in ("number", "name") if c not in col_index]
    if missing:
        return ParseResult(
            valid_rows=[],
            errors=[RowError(row=1, field=c, msg="missing required column") for c in missing],
        )

    def get(values: list, col: str) -> object:
        idx = col_index.get(col)
        if idx is None or idx >= len(values):
            return None
        return values[idx]

    valid_rows: list[RosterRow] = []
    errors: list[RowError] = []
    seen_numbers: dict[int, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in row]
        if all(_blank(v) for v in values):
            continue

        row_ok = True

        raw_number = get(values, "number")
        number: int | None = None
        if _blank(raw_number):
            errors.append(RowError(row=row_idx, field="number", msg="required"))
            row_ok = False
        else:
            try:
                number = int(raw_number)  # type: ignore[arg-type]
            except (TypeError, ValueError):
                errors.append(RowError(row=row_idx, field="number", msg="must be an integer"))
                row_ok = False
            else:
                if not (0 <= number <= 99):
                    errors.append(RowError(row=row_idx, field="number", msg="must be 0-99"))
                    row_ok = False

        raw_name = get(values, "name")
        name = str(raw_name).strip() if not _blank(raw_name) else ""
        if not name:
            errors.append(RowError(row=row_idx, field="name", msg="required"))
            row_ok = False

        raw_positions = get(values, "positions")
        positions = None if _blank(raw_positions) else str(raw_positions).strip()

        bats: str | None = None
        throws: str | None = None
        raw_bats_throws = get(values, "bats_throws")
        if not _blank(raw_bats_throws):
            parts = str(raw_bats_throws).strip().split("/")
            if len(parts) != 2:
                errors.append(
                    RowError(row=row_idx, field="bats_throws", msg="expected format 'B/T', e.g. 'R/R'"),
                )
                row_ok = False
            else:
                bats, throws = parts[0].strip().upper(), parts[1].strip().upper()
                if bats not in _BATS_VALUES:
                    errors.append(
                        RowError(row=row_idx, field="bats_throws", msg=f"bats must be one of {sorted(_BATS_VALUES)}"),
                    )
                    row_ok = False
                if throws not in _THROWS_VALUES:
                    errors.append(
                        RowError(row=row_idx, field="bats_throws", msg=f"throws must be one of {sorted(_THROWS_VALUES)}"),
                    )
                    row_ok = False

        raw_title = get(values, "title")
        title = "member"
        if not _blank(raw_title):
            raw_title_str = str(raw_title).strip()
            if raw_title_str in _TITLE_LABELS:
                title = _TITLE_LABELS[raw_title_str]
            elif raw_title_str in _TITLE_VALUES:
                title = raw_title_str
            else:
                errors.append(
                    RowError(
                        row=row_idx, field="title",
                        msg=f"must be one of {sorted(_TITLE_LABELS)}",
                    ),
                )
                row_ok = False

        birthdate: date | None = None
        raw_birthdate = get(values, "birthdate")
        if not _blank(raw_birthdate):
            birthdate, err = _parse_birthdate(raw_birthdate)
            if err:
                errors.append(RowError(row=row_idx, field="birthdate", msg=err))
                row_ok = False

        national_id: str | None = None
        raw_national_id = get(values, "national_id")
        if not _blank(raw_national_id):
            national_id = str(raw_national_id).strip().upper()
            if not _NATIONAL_ID_RE.match(national_id):
                errors.append(
                    RowError(row=row_idx, field="national_id", msg="expected 1 letter + 9 digits"),
                )
                row_ok = False

        email: str | None = None
        raw_email = get(values, "email")
        if not _blank(raw_email):
            email = str(raw_email).strip()
            if not _EMAIL_RE.match(email):
                errors.append(RowError(row=row_idx, field="email", msg="invalid email format"))
                row_ok = False

        phone: str | None = None
        raw_phone = get(values, "phone")
        if not _blank(raw_phone):
            phone = str(raw_phone).strip()

        if number is not None:
            if number in seen_numbers:
                errors.append(
                    RowError(
                        row=row_idx, field="number",
                        msg=f"duplicate number in file (also row {seen_numbers[number]})",
                    ),
                )
                row_ok = False
            else:
                seen_numbers[number] = row_idx

        if row_ok:
            valid_rows.append(
                RosterRow(
                    row=row_idx, number=number, name=name, positions=positions,  # type: ignore[arg-type]
                    bats=bats, throws=throws, title=title, birthdate=birthdate,
                    national_id=national_id, email=email, phone=phone,
                ),
            )

    return ParseResult(valid_rows=valid_rows, errors=errors)
