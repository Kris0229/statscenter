"""Excel schedule (fixture list) import: parsing + row validation only.

Mirrors app/services/roster_import.py's pure-parse-then-commit split — this
module never touches the database. Team-name -> team_id resolution happens
in the router (app/api/schedule_import.py), same as roster_import's
existing-number collision check.
"""
import io
from dataclasses import dataclass
from datetime import date, datetime, time

import openpyxl

TEMPLATE_COLUMNS = ["match_no", "date", "time", "venue", "away_team", "home_team"]


@dataclass
class ScheduleRow:
    row: int
    match_no: str | None
    game_date: date
    start_time: time | None
    venue: str | None
    away_team_name: str
    home_team_name: str


@dataclass
class RowError:
    row: int
    field: str
    msg: str


@dataclass
class ParseResult:
    valid_rows: list[ScheduleRow]
    errors: list[RowError]


def build_template_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "schedule"
    ws.append(TEMPLATE_COLUMNS)
    return wb


def _blank(value: object) -> bool:
    return value is None or str(value).strip() == ""


def _parse_date(raw: object) -> tuple[date | None, str | None]:
    if isinstance(raw, datetime):
        return raw.date(), None
    if isinstance(raw, date):
        return raw, None
    try:
        return date.fromisoformat(str(raw).strip()), None
    except ValueError:
        return None, "expected a date cell or 'YYYY-MM-DD'"


def _parse_time(raw: object) -> tuple[time | None, str | None]:
    if isinstance(raw, datetime):
        return raw.time(), None
    if isinstance(raw, time):
        return raw, None
    try:
        return time.fromisoformat(str(raw).strip()), None
    except ValueError:
        return None, "expected a time cell or 'HH:MM'"


def parse_schedule_workbook(content: bytes) -> ParseResult:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb["schedule"] if "schedule" in wb.sheetnames else wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1), ())
    header = [str(c.value).strip().lower() if c.value is not None else "" for c in header_row]
    col_index = {name: i for i, name in enumerate(header)}

    missing = [c for c in ("date", "away_team", "home_team") if c not in col_index]
    if missing:
        return ParseResult(
            valid_rows=[],
            errors=[RowError(row=1, field=c, msg="missing required column") for c in missing],
        )

    def get(values: list, col: str) -> object:
        idx = col_index.get(col)
        if idx is None or idx >= len(values):
            return None
        return values[idx]

    valid_rows: list[ScheduleRow] = []
    errors: list[RowError] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in row]
        if all(_blank(v) for v in values):
            continue

        row_ok = True

        raw_date = get(values, "date")
        game_date: date | None = None
        if _blank(raw_date):
            errors.append(RowError(row=row_idx, field="date", msg="required"))
            row_ok = False
        else:
            game_date, err = _parse_date(raw_date)
            if err:
                errors.append(RowError(row=row_idx, field="date", msg=err))
                row_ok = False

        start_time: time | None = None
        raw_time = get(values, "time")
        if not _blank(raw_time):
            start_time, err = _parse_time(raw_time)
            if err:
                errors.append(RowError(row=row_idx, field="time", msg=err))
                row_ok = False

        raw_venue = get(values, "venue")
        venue = None if _blank(raw_venue) else str(raw_venue).strip()

        raw_match_no = get(values, "match_no")
        match_no = None if _blank(raw_match_no) else str(raw_match_no).strip()

        raw_away = get(values, "away_team")
        away_team_name = str(raw_away).strip() if not _blank(raw_away) else ""
        if not away_team_name:
            errors.append(RowError(row=row_idx, field="away_team", msg="required"))
            row_ok = False

        raw_home = get(values, "home_team")
        home_team_name = str(raw_home).strip() if not _blank(raw_home) else ""
        if not home_team_name:
            errors.append(RowError(row=row_idx, field="home_team", msg="required"))
            row_ok = False

        if away_team_name and home_team_name and away_team_name == home_team_name:
            errors.append(
                RowError(row=row_idx, field="home_team", msg="away_team and home_team must differ"),
            )
            row_ok = False

        if row_ok:
            valid_rows.append(
                ScheduleRow(
                    row=row_idx, match_no=match_no, game_date=game_date,  # type: ignore[arg-type]
                    start_time=start_time, venue=venue,
                    away_team_name=away_team_name, home_team_name=home_team_name,
                ),
            )

    return ParseResult(valid_rows=valid_rows, errors=errors)
